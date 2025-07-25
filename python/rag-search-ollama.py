"""
uv add \
  faiss-cpu>=1.11.0 \
  fastapi>=0.115.14 \
  langchain>=0.3.26 \
  langchain-community>=0.3.27 \
  langchain-ollama>=0.3.3 \
  ollama>=0.5.1 \
  openpyxl>=3.1.5 \
  pandas>=2.3.0 \
  pymupdf>=1.26.3 \
  pypdf>=5.7.0 \
  python-docx>=1.2.0 \
  ruff>=0.12.2 \
  tiktoken>=0.9.0 \
  unstructured>=0.18.3 \
  uvicorn>=0.35.0
"""

import os
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from langchain.schema import Document
from langchain_community.vectorstores import FAISS
from langchain.text_splitter import CharacterTextSplitter
from langchain.chains import RetrievalQA
from langchain_community.document_loaders import (
    TextLoader,
    DirectoryLoader,
    PyPDFLoader,
    UnstructuredWordDocumentLoader,
)
from langchain_ollama import OllamaLLM, OllamaEmbeddings
from openpyxl import load_workbook


# 设置 Ollama 的主机地址
OLLAMA_HOST = "http://xxx:11434"
os.environ["OLLAMA_HOST"] = OLLAMA_HOST
app = FastAPI(title="RAG QA API")


# 文档目录
DOCS_DIR = "docs"
INDEX_DIR = "FAISS_INDEX"


def load_excel():
    # 加载 excel 文件
    filepath = "xx.xlsx"
    sheets = ["0307", "0328更新"]
    data = []
    for sheet in sheets:
        d = read_excel_fill_merged("docs/" + filepath, sheet)
        data.extend(d)
    return rows_to_documents(data=data, sheet_name=filepath)


def rows_to_documents(data, sheet_name):
    documents = []
    headers = data[0]  # 第一行是标题

    for idx, row in enumerate(data[1:], start=2):  # 从第二行开始，行号从2起
        # 拼接成“标题：内容；标题：内容；...”格式的字符串
        parts = []
        for header, cell in zip(headers, row):
            if cell is None:
                cell = ""
            parts.append(f"{header}：{cell}")
        text = "；".join(parts)
        metadata = {"sheet": sheet_name, "row": idx}
        documents.append(Document(page_content=text, metadata=metadata))

    return documents


def read_excel_fill_merged(filepath, sheet_name):
    wb = load_workbook(filename=filepath, data_only=True)
    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(values_only=True):
        # 把None替换成''，转换成列表
        clean_row = [cell if cell is not None else "" for cell in row]
        data.append(clean_row)

    merged_cells = ws.merged_cells.ranges

    for merged_range in merged_cells:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row - 1,
            merged_range.min_col - 1,
            merged_range.max_row - 1,
            merged_range.max_col - 1,
        )
        fill_value = data[min_row][min_col]
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                data[r][c] = fill_value

    return data


def load_all_docs(directory):
    all_docs = []

    # 加载 txt 文件
    txt_loader = DirectoryLoader(directory, glob="**/*.txt", loader_cls=TextLoader)
    all_docs.extend(txt_loader.load())

    # 加载 pdf 文件
    pdf_loader = DirectoryLoader(directory, glob="**/*.pdf", loader_cls=PyPDFLoader)
    all_docs.extend(pdf_loader.load())

    # 加载 docx 文件
    word_loader = DirectoryLoader(
        directory, glob="**/*.docx", loader_cls=UnstructuredWordDocumentLoader
    )
    all_docs.extend(word_loader.load())

    # 加载 excel 文件
    all_docs.extend(load_excel())
    return all_docs


# 1. 加载文档
# loader = TextLoader("docs/my_knowledge.txt", encoding="utf-8")
# documents = loader.load()
# loader = DirectoryLoader(DOCS_DIR, glob="**/*.txt", loader_cls=TextLoader)

documents = load_all_docs(DOCS_DIR)

# 2. 分割文档
text_splitter = CharacterTextSplitter(chunk_size=500, chunk_overlap=50)
docs = text_splitter.split_documents(documents)

# 3. 创建 embedding 并构建向量索引
# model: bge-m3, nomic-embed-text
embedding = OllamaEmbeddings(model="bge-m3", base_url=OLLAMA_HOST)
vectorstore = FAISS.from_documents(docs, embedding)
# 保存到硬盘
# vectorstore.save_local(INDEX_DIR)

# 4. 初始化 LLM
# model: deepseek-r1:1.5b,  qwen2.5:latest deepseek-r1:8b
llm = OllamaLLM(model="qwen2.5:latest", base_url=OLLAMA_HOST)

# 5. 创建 RAG QA Chain
qa_chain = RetrievalQA.from_chain_type(
    llm=llm, retriever=vectorstore.as_retriever(), return_source_documents=True
)

# 6. 用户查询
# query = input("请输入你的问题：\n> ")

# 7. 获取回答
# result = qa_chain.invoke({"query": "说一说多媒体"})

# 8. 打印结果
# print("\n🧠 回答：\n", result["result"])

# print("\n📚 相关片段：")
# for i, doc in enumerate(result["source_documents"]):
#     print(f"\n--- 片段 {i+1} ---\n{doc.page_content}")


# 允许的跨域来源列表，* 代表所有
origins = [
    "*",  # 允许所有域名跨域，生产环境可改成指定域名
    # "http://localhost",
    # "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  # 允许访问的域名列表
    allow_credentials=True,  # 是否允许携带Cookie等凭证
    allow_methods=["*"],  # 允许所有请求方法 GET POST PUT DELETE...
    allow_headers=["*"],  # 允许所有请求头
)


class QueryRequest(BaseModel):
    query: str


@app.post("/query")
async def query_qa(req: QueryRequest):
    if not req.query.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty")

    result = qa_chain.invoke({"query": req.query})
    return {
        "answer": result["result"],
        # "source_documents": [
        #     {"page_content": doc.page_content} for doc in result["source_documents"]
        # ],
        # "result": result,
    }
